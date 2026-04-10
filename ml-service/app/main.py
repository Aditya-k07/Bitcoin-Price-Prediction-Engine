"""
CoinSight ML Service — FastAPI Application

Serves Bitcoin price predictions via REST API.
Endpoints:
  GET  /predict?model=xgboost|prophet&days=30
  GET  /predict/export?model=xgboost|lstm_xgboost&days=30
  POST /retrain?model=xgboost|prophet
  GET  /health
"""

import logging
import os
import threading
import io
from contextlib import asynccontextmanager
from datetime import datetime

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.data_loader import load_daily_data, load_from_coingecko_ohlc
from app.features import engineer_features, get_feature_columns
from app.models.xgboost_model import XGBoostPredictor
from app.models.ridge_model import RidgePredictor
from app.schemas import PredictionResponse, RetrainResponse, HealthResponse, HistoricalDataRequest

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger(__name__)

# Global model instances
xgboost_predictor = XGBoostPredictor()
ridge_predictor = RidgePredictor()

# Global data cache
_daily_data = None
_featured_data = None
_latest_featured_data = None
_latest_signature = None
_data_lock = threading.Lock()


def _load_data():
    """Load and cache the dataset."""
    global _daily_data, _featured_data
    with _data_lock:
        if _daily_data is None:
            csv_path = os.environ.get("BTC_DATA_PATH", None)
            _daily_data = load_daily_data(csv_path)
            _featured_data = engineer_features(_daily_data)
    return _daily_data, _featured_data


def _get_data():
    """Get cached data, loading if necessary."""
    if _daily_data is None:
        return _load_data()
    return _daily_data, _featured_data


def _try_load_models():
    """Try to load pre-trained models from disk on startup."""
    xgb_loaded = xgboost_predictor.load()
    ridge_loaded = ridge_predictor.load()

    if xgb_loaded:
        logger.info("XGBoost model loaded from saved state.")
    if ridge_loaded:
        logger.info("Ridge model loaded from saved state.")

    return xgb_loaded, ridge_loaded


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application startup and shutdown lifecycle."""
    logger.info("🚀 CoinSight ML Service starting up...")

    # Try loading saved models (non-blocking — data loading happens on first request)
    _try_load_models()

    logger.info("✅ ML Service ready.")
    yield
    logger.info("🛑 ML Service shutting down.")


# Create FastAPI app
app = FastAPI(
    title="CoinSight ML Service",
    description="Bitcoin price prediction using XGBoost and Prophet",
    version="1.0.0",
    lifespan=lifespan,
)

# CORS middleware for frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health", response_model=HealthResponse)
async def health_check():
    """
    Health check endpoint.
    Returns service status and which models are loaded.
    """
    models_loaded = []
    if xgboost_predictor.is_trained:
        models_loaded.append("xgboost")
    if ridge_predictor.is_trained:
        models_loaded.append("ridge")

    return HealthResponse(
        status="healthy",
        models_loaded=models_loaded
    )


@app.get("/predict", response_model=PredictionResponse)
async def predict(
    model: str = Query("xgboost", regex="^(xgboost|ridge)$",
                       description="Model to use: 'xgboost' (accurate) or 'ridge' (fast)"),
    days: int = Query(30, ge=1, le=365,
                      description="Number of days to predict (1-365)")
):
    """
    Generate price predictions with confidence intervals.

    - **model**: 'xgboost' (accurate, slower) or 'ridge' (fast, lightweight)
    - **days**: Number of future days to predict (default: 30, max: 365)

    Returns predicted prices with 95% confidence bounds.
    """
    daily_data, featured_data = _get_data()

    if model == "xgboost":
        predictor = xgboost_predictor
        if not predictor.is_trained:
            # Auto-train on first request
            logger.info("XGBoost not trained, training now...")
            feature_cols = get_feature_columns(featured_data)
            predictor.train(featured_data, feature_cols)

        predictions = predictor.predict(featured_data, days=days)
        metrics = predictor.metrics

    elif model == "ridge":
        predictor = ridge_predictor
        if not predictor.is_trained:
            logger.info("Ridge not trained, training now...")
            feature_cols = get_feature_columns(featured_data)
            predictor.train(featured_data, feature_cols)

        predictions = predictor.predict(featured_data, days=days)
        metrics = predictor.metrics

    else:
        raise HTTPException(status_code=400, detail=f"Unknown model: {model}")

    logger.info(f"Raw predictions from model: {predictions[:2] if len(predictions) > 0 else 'EMPTY'}")
    
    # Convert predictions to OHLC format
    ohlc_predictions = []
    for pred in predictions:
        ohlc_predictions.append({
            "date": pred["date"],
            "open": pred["price"],
            "high": pred["upper"],
            "low": pred["lower"],
            "close": pred["price"]
        })
    
    logger.info(f"Converted OHLC predictions: {ohlc_predictions[:2] if len(ohlc_predictions) > 0 else 'EMPTY'}")

    return PredictionResponse(
        model=model,
        predictions=ohlc_predictions,
        rmse=metrics.rmse,
        mae=metrics.mae,
        r2_score=metrics.r2_score,
        mape=metrics.mape,
        f1_score=metrics.f1_score,
        accuracy=metrics.accuracy,
        directional_accuracy=metrics.directional_accuracy,
        trained_at=metrics.trained_at.isoformat() + "Z",
        architecture_details=metrics.architecture_details
    )


@app.post("/predict_with_data", response_model=PredictionResponse)
async def predict_with_data(request: HistoricalDataRequest):
    """
    Generate predictions using fresh OHLC data from the Go backend.

    This endpoint accepts OHLC candlestick data (from CoinGecko via Go backend),
    trains/uses a model, and returns predictions with confidence intervals.

    - **model**: 'xgboost' (accurate) or 'ridge' (fast)
    - **days**: Number of future days to predict (1-365)
    - **data**: List of OHLC candles with timestamp, open, high, low, close

    Returns predicted prices with 95% confidence bounds.
    """
    if request.model not in ["xgboost", "ridge"]:
        raise HTTPException(status_code=400, detail=f"Unknown model: {request.model}")

    logger.info(f"[/predict_with_data] Received {len(request.data)} candles, predicting {request.days} days")

    global _latest_featured_data, _latest_signature

    try:
        # Convert received OHLC data to DataFrame
        candle_dicts = [
            {
                "timestamp": c.timestamp,
                "open": c.open,
                "high": c.high,
                "low": c.low,
                "close": c.close
            }
            for c in request.data
        ]
        daily_data = load_from_coingecko_ohlc(candle_dicts)
        featured_data = engineer_features(daily_data)
        _latest_featured_data = featured_data.copy()
        latest_ts = int(candle_dicts[-1]["timestamp"]) if candle_dicts else 0
        data_signature = (request.model, len(candle_dicts), latest_ts)

    except Exception as e:
        logger.error(f"Failed to process OHLC data: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to process data: {str(e)}")

    if request.model == "xgboost":
        predictor = xgboost_predictor
        feature_cols = get_feature_columns(featured_data)
        # Retrain when cold, CoinGecko slice changed, or caller explicitly retrains.
        if (
            request.force_retrain
            or (not predictor.is_trained)
            or (_latest_signature != data_signature)
        ):
            logger.info("Training XGBoost on provided data...")
            predictor.train(featured_data, feature_cols)
            _latest_signature = data_signature

        predictions = predictor.predict(featured_data, days=request.days)
        metrics = predictor.metrics

    elif request.model == "ridge":
        predictor = ridge_predictor
        feature_cols = get_feature_columns(featured_data)
        if (
            request.force_retrain
            or (not predictor.is_trained)
            or (_latest_signature != data_signature)
        ):
            logger.info("Training Ridge on provided data...")
            predictor.train(featured_data, feature_cols)
            _latest_signature = data_signature

        predictions = predictor.predict(featured_data, days=request.days)
        metrics = predictor.metrics

    logger.info(f"Raw predictions from model: {predictions[:2] if len(predictions) > 0 else 'EMPTY'}")

    # Convert predictions to OHLC format
    ohlc_predictions = []
    for pred in predictions:
        ohlc_predictions.append({
            "date": pred["date"],
            "open": pred["price"],
            "high": pred["upper"],
            "low": pred["lower"],
            "close": pred["price"]
        })

    logger.info(f"Converted OHLC predictions: {ohlc_predictions[:2] if len(ohlc_predictions) > 0 else 'EMPTY'}")

    return PredictionResponse(
        model=request.model,
        predictions=ohlc_predictions,
        rmse=metrics.rmse,
        mae=metrics.mae,
        r2_score=metrics.r2_score,
        mape=metrics.mape,
        f1_score=metrics.f1_score,
        accuracy=metrics.accuracy,
        directional_accuracy=metrics.directional_accuracy,
        trained_at=metrics.trained_at.isoformat() + "Z",
        architecture_details=metrics.architecture_details
    )


@app.post("/retrain", response_model=RetrainResponse)
async def retrain(
    model: str = Query("xgboost", regex="^(xgboost|ridge)$",
                       description="Model to retrain: 'xgboost' or 'ridge'")
):
    """
    Retrain a model and return updated metrics.

    This endpoint reloads the data, re-engineers features, and retrains
    the specified model from scratch. Useful when new data is available
    or when you want to refresh the model.
    """
    global _daily_data, _featured_data

    # Force reload data
    logger.info(f"Retraining {model} model...")
    with _data_lock:
        csv_path = os.environ.get("BTC_DATA_PATH", None)
        _daily_data = load_daily_data(csv_path)
        _featured_data = engineer_features(_daily_data)

    if model == "xgboost":
        feature_cols = get_feature_columns(_featured_data)
        metrics = xgboost_predictor.train(_featured_data, feature_cols)
    elif model == "ridge":
        feature_cols = get_feature_columns(_featured_data)
        metrics = ridge_predictor.train(_featured_data, feature_cols)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown model: {model}")

    return RetrainResponse(
        model=model,
        rmse=metrics.rmse,
        mae=metrics.mae,
        r2_score=metrics.r2_score,
        mape=metrics.mape,
        f1_score=metrics.f1_score,
        accuracy=metrics.accuracy,
        directional_accuracy=metrics.directional_accuracy,
        trained_at=metrics.trained_at.isoformat() + "Z",
        message=f"{model.capitalize()} model retrained successfully",
        architecture_details=metrics.architecture_details
    )


@app.get("/predict/export")
async def export_predictions_to_excel(
    model: str = Query("xgboost", regex="^(xgboost|ridge)$"),
    days: int = Query(30, ge=1, le=365)
):
    """
    Export predictions to Excel format with date and predicted price.
    
    Returns an Excel file (.xlsx) that can be downloaded.
    Contains columns: Date, Open Price, High (Upper Bound), Low (Lower Bound), Close Price
    """
    global _latest_featured_data

    # Prefer dataset-backed data, but gracefully fall back to the latest
    # feature-engineered market data received via /predict_with_data.
    try:
        _, featured_data = _get_data()
    except Exception:
        if _latest_featured_data is None:
            raise HTTPException(
                status_code=503,
                detail="No historical dataset available for export. Generate predictions first."
            )
        featured_data = _latest_featured_data.copy()

    if model == "xgboost":
        predictor = xgboost_predictor
        if not predictor.is_trained:
            logger.info("XGBoost not trained, training now...")
            feature_cols = get_feature_columns(featured_data)
            predictor.train(featured_data, feature_cols)
        predictions = predictor.predict(featured_data, days=days)
    elif model == "ridge":
        predictor = ridge_predictor
        if not predictor.is_trained:
            logger.info("Ridge not trained, training now...")
            feature_cols = get_feature_columns(featured_data)
            predictor.train(featured_data, feature_cols)
        predictions = predictor.predict(featured_data, days=days)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown model: {model}")

    # Create DataFrame from predictions
    df = pd.DataFrame([
        {
            "Date": pred["date"],
            "Predicted Price (USD)": pred["price"],
            "Upper Bound (95% CI)": pred["upper"],
            "Lower Bound (95% CI)": pred["lower"],
            "Price Range": pred["upper"] - pred["lower"]
        }
        for pred in predictions
    ])

    # Create Excel workbook
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Predictions')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Predictions']
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Format data cells and set column widths
        column_widths = [15, 20, 20, 20, 15]
        number_format = '0.00'
        
        for idx, column in enumerate(worksheet.columns, 1):
            worksheet.column_dimensions[column[0].column_letter].width = column_widths[idx - 1]
            
            for cell in column[1:]:  # Skip header
                cell.alignment = data_alignment
                cell.border = border
                if idx > 1:  # Format price columns as currency
                    cell.number_format = number_format
        
        # Add metadata sheet
        metadata_ws = workbook.create_sheet('Metadata')
        metadata_ws['A1'] = "Bitcoin Price Prediction Report"
        metadata_ws['A1'].font = Font(bold=True, size=14)
        metadata_ws['A3'] = "Model Used:"
        metadata_ws['B3'] = model
        metadata_ws['A4'] = "Prediction Days:"
        metadata_ws['B4'] = days
        metadata_ws['A5'] = "Generated At:"
        metadata_ws['B5'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        
        metadata_ws.column_dimensions['A'].width = 20
        metadata_ws.column_dimensions['B'].width = 30
    
    excel_buffer.seek(0)
    
    filename = f"bitcoin_predictions_{model}_{days}days_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    logger.info(f"Generated Excel export: {filename}")
    
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("app.main:app", host="0.0.0.0", port=port, reload=True)
